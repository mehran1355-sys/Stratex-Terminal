"""
تولید گزارش‌های اکسل
Excel Reporter Module
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReporter:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_analysis_report(self, timeframe: str, analyses: List[Dict], filename: str = None) -> str:
        if not analyses:
            return ""
        if filename is None:
            filename = f"Analysis_{timeframe}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        filepath = self.output_dir / filename
        data = []
        for a in analyses:
            entries = a.get("entry_zones", {})
            entry_str = ", ".join([f"{v:.4f}" for v in entries.values()]) if isinstance(entries, dict) else str(entries)
            data.append({"نماد": a.get("symbol", ""), "تایم‌فریم": timeframe, "جهت": a.get("direction", ""), "نوع شکست": a.get("break_type", ""), "مناطق ورود": entry_str, "TP1": round(a.get("tp1", 0), 5), "TP2": round(a.get("tp2", 0), 5), "SL": round(a.get("sl", 0), 5), "حجم": a.get("volume", 0), "وضعیت": a.get("status", "در انتظار")})
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        logger.info(f"Report saved: {filepath}")
        return str(filepath)

    def create_multi_timeframe_report(self, analyses_by_tf: Dict[str, List[Dict]], filename: str = None) -> str:
        if filename is None:
            filename = f"Full_Analysis_{datetime.now():%Y%m%d_%H%M}.xlsx"
        filepath = self.output_dir / filename
        with pd.ExcelWriter(filepath) as writer:
            for tf, analyses in analyses_by_tf.items():
                if analyses:
                    df = pd.DataFrame(analyses)
                    df.to_excel(writer, sheet_name=f"Analysis_{tf}", index=False)
        logger.info(f"Multi-TF report saved: {filepath}")
        return str(filepath)

    def create_trade_history_report(self, trades: List[Dict], period: str = "daily") -> str:
        filename = f"Trade_History_{period}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        filepath = self.output_dir / filename
        data = []
        for t in trades:
            data.append({"شماره": t.get("ticket", ""), "نماد": t.get("symbol", ""), "جهت": t.get("direction", ""), "ورود": t.get("entry_price", 0), "خروج": t.get("exit_price", 0), "حجم": t.get("volume", 0), "سود/ضرر": t.get("pnl", 0), "دلیل": t.get("exit_reason", ""), "تاریخ": t.get("entry_time", "")})
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        return str(filepath)

    def create_backtest_report(self, result, charts: Dict = None) -> str:
        filename = f"Backtest_{result.config.symbol}_{result.config.timeframe}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        filepath = self.output_dir / filename
        if OPENPYXL_AVAILABLE:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "خلاصه"
            summary_data = [("نماد", result.config.symbol), ("تایم‌فریم", result.config.timeframe), ("بازه", f"{result.config.start_date.date()} تا {result.config.end_date.date()}"), ("کل معاملات", result.total_trades), ("Win Rate", f"{result.win_rate:.1f}%"), ("سود/ضرر کل", f"${result.total_pnl:,.2f}"), ("Profit Factor", f"{result.profit_factor:.2f}"), ("Sharpe Ratio", f"{result.sharpe_ratio:.2f}"), ("Max Drawdown", f"${result.max_drawdown:,.2f} ({result.max_drawdown_pct:.2f}%)")]
            for i, (k, v) in enumerate(summary_data, 2):
                ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
                ws1.cell(row=i, column=2, value=v)
            wb.save(filepath)
        else:
            pd.DataFrame({"metric": [k for k, _ in summary_data], "value": [v for _, v in summary_data]}).to_excel(filepath, index=False)
        return str(filepath)

    def get_latest_report(self, report_type: str = "analysis") -> Optional[str]:
        files = list(self.output_dir.glob(f"*{report_type}*.xlsx"))
        if files:
            return str(max(files, key=lambda f: f.stat().st_ctime))
        return None
